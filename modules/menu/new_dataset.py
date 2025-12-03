def set_window_icon(window):
    """Helper function to set the custom icon for any tkinter window"""
    try:
        icon_path = os.path.join(os.path.dirname(__file__), "MTT_favicon.png")
        if os.path.exists(icon_path):
            from PIL import Image, ImageTk
            icon_image = Image.open(icon_path)
            icon_photo = ImageTk.PhotoImage(icon_image)
            window.iconphoto(True, icon_photo)
    except Exception as e:
        print(f"Could not load icon: {e}")

def new_dataset(root=None):
    # Use provided root or create new one
    if root is None:
        newdatamenu= tk.Tk()
    else:
        newdatamenu = root

    # Create the main window
    newdatamenu.title("FAIR.rdm MTT Beta 1.0")
    newdatamenu.geometry("800x750")

    def refresh_treeview():
        """Aktualisiert die Treeview mit den aktuellen Werten aus df_current_data."""
        tree.delete(*tree.get_children())
        if df_current_data is not None:
            for index, row in df_current_data.iterrows():
                tree.insert('', tk.END, values=(row['Metadata Property'], row['Metadata Value']))
        else:
            tree.insert('', tk.END, values=("No data", "Please create a dataset first"))
    
    def save_to_excel_with_formatting(property_name, new_value):
        """Gemeinsame Funktion zum Speichern in Excel mit Formatierung"""
        try:
            from openpyxl import load_workbook
            import copy
            
            wb = load_workbook(current_file_path)
            ws = wb.active
            
            # Find and update the row
            for row in ws.iter_rows():
                if row[0].value == property_name:
                    # Store original formatting
                    original_font = copy.copy(row[1].font)
                    original_fill = copy.copy(row[1].fill)
                    original_alignment = copy.copy(row[1].alignment)
                    original_border = copy.copy(row[1].border)
                    original_number_format = row[1].number_format
                    
                    # Update value
                    row[1].value = new_value
                    
                    # Restore formatting
                    row[1].font = original_font
                    row[1].fill = original_fill
                    row[1].alignment = original_alignment
                    row[1].border = original_border
                    row[1].number_format = original_number_format
                    break
            
            wb.save(current_file_path)
            print(f"Updated {property_name} to: {new_value}")
            return True
            
        except Exception as e:
            print(f"Error saving {property_name}: {e}")
            return False

    def batch_save_to_excel_with_formatting(updates_dict):
        """Batch-Update für mehrere Felder gleichzeitig"""
        try:
            from openpyxl import load_workbook
            import copy
            
            wb = load_workbook(current_file_path)
            ws = wb.active
            
            # Update all fields in one go
            for row in ws.iter_rows():
                property_name = row[0].value
                if property_name in updates_dict:
                    new_value = updates_dict[property_name]
                    
                    # Store original formatting
                    original_font = copy.copy(row[1].font)
                    original_fill = copy.copy(row[1].fill)
                    original_alignment = copy.copy(row[1].alignment)
                    original_border = copy.copy(row[1].border)
                    original_number_format = row[1].number_format
                    
                    # Update value
                    row[1].value = new_value
                    
                    # Restore formatting
                    row[1].font = original_font
                    row[1].fill = original_fill
                    row[1].alignment = original_alignment
                    row[1].border = original_border
                    row[1].number_format = original_number_format
            
            wb.save(current_file_path)
            print(f"Batch updated {len(updates_dict)} fields")
            return True
            
        except Exception as e:
            print(f"Error in batch save: {e}")
            return False
    
    def retrieve_data():
        """Retrieve data from the repository and update the DataFrame and Treeview"""
        nonlocal df_current_data
        print("Retrieve data button clicked")
        # Erstelle ein nicht-blockierendes "Bitte warten"-Fenster
        wait_win = tk.Toplevel(newdatamenu)
        wait_win.title("Bitte warten")
        wait_win.geometry("250x80")
        wait_win.transient(newdatamenu)
        wait_win.grab_set()
        tk.Label(wait_win, text="Retrieving data...\nPlease wait.", font=("Helvetica", 11)).pack(expand=True, pady=15)
        wait_win.update()
        try:
            #get the dataset_uri from df
            df_data = df_current_data.fillna('not_defined')
            has_landing_page = df_data.loc[df_data['Metadata Property'] == 'has_landing_page', 'Metadata Value'].values
            has_landing_page = has_landing_page[0]
            if has_landing_page == 'not_defined':
                messagebox.showerror("Error", "No landing page defined in this dataset.")
            else:
                uri = has_landing_page
                module_to_use = rec_repo(uri)
                print(f"Using module: {module_to_use}")
                #now execute all functions from the recognized module
                if module_to_use != "not_recognized":
                    # Dynamischer Import: modules.modulname
                    module = importlib.import_module(module_to_use)
                    local_file = getattr(module, 'local_file', None)
                    if local_file == True:
                        #Open a dialog to select local file
                        from tkinter import filedialog
                        file_path = filedialog.askopenfilename(title="Select local file for metadata retrieval")
                        if not file_path:
                            messagebox.showerror("Error", "No file selected for local metadata retrieval.")
                            return
                        uri = file_path
                    else:
                        uri = has_landing_page
                    data = module.get_metadata(uri)

                    # Dictionary für Batch-Updates
                    updates_to_save = {}
                    
                    #print(data)
                    #1. has_identifier
                    #2. has_title
                    has_title = module.get_title(data)
                    if has_title != 'not_defined':
                        df_data.loc[df_data['Metadata Property'] == 'has_title', 'Metadata Value'] = has_title
                        updates_to_save['has_title'] = has_title
                        print(f"Title: {has_title}")
                    
                    #3. has_description
                    has_description = module.get_description(data)
                    if has_description != 'not_defined':
                        df_data.loc[df_data['Metadata Property'] == 'has_description', 'Metadata Value'] = has_description
                        updates_to_save['has_description'] = has_description
                        print(f"Description: {has_description}")
                    
                    #4. was_issued
                    was_issued = module.get_was_issued(data)
                    if was_issued != 'not_defined':
                        df_data.loc[df_data['Metadata Property'] == 'was_issued', 'Metadata Value'] = was_issued
                        updates_to_save['was_issued'] = was_issued
                        print(f"Was Issued: {was_issued}")
                    
                    #5. was_modified
                    was_modified = module.get_was_modified(data)
                    if was_modified != 'not_defined':
                        df_data.loc[df_data['Metadata Property'] == 'was_modified', 'Metadata Value'] = was_modified
                        updates_to_save['was_modified'] = was_modified
                        print(f"Was Modified: {was_modified}")
                    
                    #6. has_publisher
                    has_publisher = module.get_publisher(data)
                    if has_publisher != 'not_defined':
                        df_data.loc[df_data['Metadata Property'] == 'has_publisher', 'Metadata Value'] = has_publisher
                        updates_to_save['has_publisher'] = has_publisher
                        print(f"Publisher: {has_publisher}")
                    # Alle weiteren Felder hinzufügen
                    field_getters = [
                        ('has_contributor', 'get_contributor'),
                        ('has_creator', 'get_creator'),
                        ('has_owner', 'get_owner'),
                        ('has_responsible', 'get_responsible'),
                        ('has_original_id', 'get_original_id'),
                        ('has_ariadne_subject', 'get_ariadne_subject'),
                        ('has_native_subject', 'get_native_subject'),
                        ('has_derived_subject_uri', 'get_derived_subject_uri'),
                        ('has_derived_subject_term', 'get_derived_subject_term'),
                        ('has_language', 'get_language'),
                        ('was_created_on', 'get_was_created_on'),
                        ('has_access_policy', 'get_access_policy'),
                        ('has_access_rights', 'get_access_rights'),
                        ('has_extent', 'get_extent'),
                        ('has_periodo_uri', 'get_periodo_uri'),
                        ('has_native_period', 'get_native_period'),
                        ('has_chronontology_uri', 'get_chronontology_uri'),
                        ('has_from', 'get_from'),
                        ('has_until', 'get_until'),
                        ('has_visual_component', 'get_visual_component'),
                        ('has_part', 'get_has_part'),
                        ('is_part_of', 'get_is_part_of')
                    ]
                    
                    # Alle Standard-Felder verarbeiten
                    for field_name, getter_name in field_getters:
                        if hasattr(module, getter_name):
                            value = getattr(module, getter_name)(data)
                            if value != 'not_defined':
                                df_data.loc[df_data['Metadata Property'] == field_name, 'Metadata Value'] = value
                                updates_to_save[field_name] = value
                                print(f"{field_name}: {value}")
                    
                    # Spezielle Behandlung für räumliche Daten
                    localization_type = module.has_point(data)
                    print(f"Localization Type: {localization_type}")
                    
                    if localization_type == "is_bounding_box":
                        bb_fields = [
                            ('bb_has_place_name', 'get_bb_place_name'),
                            ('bb_has_coordinate_system', 'get_bb_coordinate_system'),
                            ('bb_has_country_code', 'get_bb_country_code'),
                            ('bb_has_place_uri', 'get_bb_place_uri'),
                            ('has_bounding_box_min_lat', 'get_bb_min_lat'),
                            ('has_bounding_box_min_lon', 'get_bb_min_lon'),
                            ('has_bounding_box_max_lat', 'get_bb_max_lat'),
                            ('has_bounding_box_max_lon', 'get_bb_max_lon')
                        ]
                        for field_name, getter_name in bb_fields:
                            if hasattr(module, getter_name):
                                value = getattr(module, getter_name)(data)
                                if value != 'not_defined':
                                    df_data.loc[df_data['Metadata Property'] == field_name, 'Metadata Value'] = value
                                    updates_to_save[field_name] = value
                                    print(f"{field_name}: {value}")
                                    
                    elif localization_type == "is_point":
                        point_fields = [
                            ('point_has_place_name', 'point_get_place_name'),
                            ('point_has_coordinate_system', 'point_get_coordinate_system'),
                            ('point_has_country_code', 'point_get_country_code'),
                            ('point_has_place_uri', 'point_get_place_uri'),
                            ('has_latitude', 'get_lat'),
                            ('has_longitude', 'get_lon')
                        ]
                        for field_name, getter_name in point_fields:
                            if hasattr(module, getter_name):
                                value = getattr(module, getter_name)(data)
                                if value != 'not_defined':
                                    df_data.loc[df_data['Metadata Property'] == field_name, 'Metadata Value'] = value
                                    updates_to_save[field_name] = value
                                    print(f"{field_name}: {value}")
                                    
                    elif localization_type == "is_polygon":
                        polygon_fields = [
                            ('polygon_has_place_name', 'polygon_has_place_name'),
                            ('polygon_has_coordinate_system', 'polygon_has_coordinate_system'),
                            ('polygon_has_country_code', 'polygon_has_country_code'),
                            ('polygon_has_place_uri', 'polygon_has_place_uri'),
                            ('has_polygonal_representation', 'get_polygonal_representation')
                        ]
                        for field_name, getter_name in polygon_fields:
                            if hasattr(module, getter_name):
                                value = getattr(module, getter_name)(data)
                                if value != 'not_defined':
                                    df_data.loc[df_data['Metadata Property'] == field_name, 'Metadata Value'] = value
                                    updates_to_save[field_name] = value
                                    print(f"{field_name}: {value}")
                    
                    # Batch-Save aller Updates mit Formatierung
                    if updates_to_save:
                        batch_save_to_excel_with_formatting(updates_to_save)
                        print(f"Saved {len(updates_to_save)} updates to Excel with preserved formatting")
                    
            df_current_data = df_data.fillna('not_defined')
            refresh_treeview()
            
            # Update missing values list after data retrieval
            update_missing_values_list()
            
        finally:
            wait_win.destroy()
        return df_current_data


    def save_edit(event=None):
        """Save the edited value"""
        nonlocal edit_item, edit_entry
        
        if edit_entry and edit_item:
            new_value = edit_entry.get()
            old_values = tree.item(edit_item, 'values')
            tree.item(edit_item, values=(old_values[0], new_value))
            
            # Update the DataFrame and save to Excel with formatting
            property_name = old_values[0]
            df_current_data.loc[df_current_data['Metadata Property'] == property_name, 'Metadata Value'] = new_value
            
            # Use the shared formatting function
            save_to_excel_with_formatting(property_name, new_value)
                
            # Refresh the missing values list after saving
            update_missing_values_list()
        
        cleanup_edit()

    def open_with_excel():
        print("Open current file with Excel")
        os.startfile(current_file_path)

    def xml_initiate_conversion():
        #Check if every mandatory field is filled
        missing_items = check_missing_values()
        if any("MANDATORY" in item for item in missing_items):
            messagebox.showerror("Error", "Cannot convert to XML. Some mandatory fields are still missing.")
            return
        else:
            print("Convert current data to XML format")
            nonlocal df_current_data
            df_current_data = df_current_data.fillna('not_defined')
            data_dict = df_current_data.set_index(df_current_data.columns[0]).T.to_dict()
            #1. Get header and footer from xml_conversion module
            header = xml_conversion.header
            footer = xml_conversion.footer
            #Building final XML data
            if "collection" in data_dict.get('Resource_Type').get('Metadata Value'):
                xmldata = f'{header}{xml_conversion.get_identifier(data_dict)}{xml_conversion.get_title(data_dict)}{xml_conversion.get_description(data_dict)}{xml_conversion.get_issued(data_dict)}{xml_conversion.get_modified(data_dict)}{xml_conversion.get_publisher(data_dict)}{xml_conversion.get_contributor(data_dict)}{xml_conversion.get_creator(data_dict)}{xml_conversion.get_owner(data_dict)}{xml_conversion.get_responsible(data_dict)}{xml_conversion.get_original_id(data_dict)}{xml_conversion.get_ariadne_subject(data_dict)}{xml_conversion.get_native_subject(data_dict)}{xml_conversion.get_derived_subject_uri(data_dict)}{xml_conversion.get_derived_subject_label(data_dict)}{xml_conversion.get_language(data_dict)}{xml_conversion.get_created_on(data_dict)}{xml_conversion.get_landing_page(data_dict)}{xml_conversion.get_access_policy(data_dict)}{xml_conversion.get_access_rights(data_dict)}{xml_conversion.get_extent(data_dict)}{xml_conversion.get_from(data_dict)}{xml_conversion.get_until(data_dict)}{xml_conversion.get_spatial_coverage(data_dict)}{xml_conversion.get_visual_component(data_dict)}{xml_conversion.get_is_part_of(data_dict)}{footer}'
            else:
                xmldata = f'{header}{xml_conversion.get_identifier(data_dict)}{xml_conversion.get_title(data_dict)}{xml_conversion.get_description(data_dict)}{xml_conversion.get_issued(data_dict)}{xml_conversion.get_modified(data_dict)}{xml_conversion.get_publisher(data_dict)}{xml_conversion.get_contributor(data_dict)}{xml_conversion.get_creator(data_dict)}{xml_conversion.get_owner(data_dict)}{xml_conversion.get_responsible(data_dict)}{xml_conversion.get_original_id(data_dict)}{xml_conversion.get_ariadne_subject(data_dict)}{xml_conversion.get_native_subject(data_dict)}{xml_conversion.get_derived_subject_uri(data_dict)}{xml_conversion.get_derived_subject_label(data_dict)}{xml_conversion.get_language(data_dict)}{xml_conversion.get_created_on(data_dict)}{xml_conversion.get_landing_page(data_dict)}{xml_conversion.get_access_policy(data_dict)}{xml_conversion.get_access_rights(data_dict)}{xml_conversion.get_extent(data_dict)}{xml_conversion.get_from(data_dict)}{xml_conversion.get_until(data_dict)}{xml_conversion.get_data_type(data_dict)}{xml_conversion.get_data_format(data_dict)}{xml_conversion.get_spatial_coverage(data_dict)}{xml_conversion.get_visual_component(data_dict)}{xml_conversion.get_is_part_of(data_dict)}{footer}'
            #print(xmldata)
            #Fortmatting the final XML data with indentation for better readability
            dom = xml.dom.minidom.parseString(xmldata)
            formatted_xml = dom.toprettyxml()
            print(formatted_xml)
            filename = data_dict.get('has_identifier').get('Metadata Value')
            #save file to metadata_mirror folder
            with open(os.path.join(os.path.dirname(__file__), f"../../metadata_mirror/{filename}.xml"), 'w', encoding='utf-8') as f:
                f.write(formatted_xml)
            messagebox.showinfo("Success", f"XML file {filename}.xml has been created successfully in metadata_mirror folder.")

    def show_xml():
        print("Opening most current XML file...")
        filename = df_current_data.loc[df_current_data['Metadata Property'] == 'has_identifier', 'Metadata Value'].values[0]
        if filename == 'not_defined':
            messagebox.showerror("Error", "No identifier defined in this dataset.")
        else:
            xml_path = os.path.abspath(os.path.join(os.path.dirname(__file__), f"../../metadata_mirror/{filename}.xml"))
            if os.path.exists(xml_path):
                os.startfile(xml_path)
            else:
                messagebox.showerror("Error", f"XML file {filename}.xml does not exist. Please convert to XML first.")

    def go_back():
        print("Button Back clicked")
        main_menu.main_menu_design(root=root)
        newdatamenu.destroy()

    newdatamenu.grid_columnconfigure(0, weight=1)
    newdatamenu.grid_columnconfigure(1, weight=1)
    newdatamenu.grid_rowconfigure(0, weight=1)
    newdatamenu.grid_rowconfigure(1, weight=1)
    newdatamenu.grid_rowconfigure(2, weight=1)
    newdatamenu.grid_rowconfigure(3, weight=1)
    newdatamenu.grid_rowconfigure(4, weight=1)
    newdatamenu.grid_rowconfigure(5, weight=1)
    newdatamenu.grid_rowconfigure(6, weight=1)
    newdatamenu.grid_rowconfigure(7, weight=1)
    newdatamenu.grid_rowconfigure(8, weight=1)

    #Titel 1. Zeile
    label_title = tk.Label(newdatamenu, text="Creation of a new Metadataset", font=("Helvetica", 16, "bold"), anchor="w")
    label_title.grid(row=0, column=0, columnspan=2, pady=10, padx=10, sticky="nw")

    #Untertitel 2. Zeile
    label_subtitle = tk.Label(newdatamenu, text="Fetch data from repository of origin and add missing values manually.", font=("Helvetica", 12, "bold"), anchor="w")
    label_subtitle.grid(row=1, column=0, columnspan=2, pady=10, padx=10, sticky="nw")

    #Button 3. Zeile
    button1 = tk.Button(newdatamenu, text="Retrieve Data", command=lambda: retrieve_data(), width=35, height=2)
    button1.grid(row=2, column=0, pady=10, padx=15, sticky="nw")

    #Button 4. Zeile
    button2 = tk.Button(newdatamenu, text="Save Progress", command=save_edit, width=35, height=2)
    button2.grid(row=3, column=0, pady=10, padx=15, sticky="nw")

    #Button 5. Zeile
    button3 = tk.Button(newdatamenu, text="Open with Excel", command=open_with_excel, width=35, height=2)
    button3.grid(row=4, column=0, pady=10, padx=15, sticky="nw")

    #Button 6. Zeile
    button4 = tk.Button(newdatamenu, text="Convert to XML", command=xml_initiate_conversion, width=35, height=2)
    button4.grid(row=5, column=0, pady=10, padx=15, sticky="nw")

    #Button 7. Zeile
    button5 = tk.Button(newdatamenu, text="Show XML", command=show_xml, width=35, height=2)
    button5.grid(row=6, column=0, pady=10, padx=15, sticky="nw")

    # Load metadata files and current data first
    metadata_tables_path = os.path.join(os.path.dirname(__file__), "../../metadata_tables")
    metadata_files = [f for f in os.listdir(metadata_tables_path) if os.path.isfile(os.path.join(metadata_tables_path, f))]
    
    # Filter for relevant files and get the most recent one
    relevant_files = [f for f in metadata_files]
    most_recent_file = max(relevant_files, key=lambda f: os.path.getmtime(os.path.join(metadata_tables_path, f)))
    current_file_path = os.path.join(metadata_tables_path, most_recent_file)
    current_file_name = os.path.splitext(os.path.basename(current_file_path))[0]
    df_current_data = pd.read_excel(current_file_path)
    #insert current identifier (which equals to the file name without extension)
    df_current_data.at[1, "Metadata Value"] = current_file_name
    #save the identifier as well to the excel file but keep it formatted
    save_to_excel_with_formatting('has_identifier', current_file_name)
    
    #ToDo Listbox in Zeile 8 - Show missing values
    scrollbar_frame = tk.Frame(newdatamenu)
    scrollbar_frame.grid(row=7, column=0, pady=10, padx=15, sticky="w")

    label_list = tk.Label(scrollbar_frame, text="Missing Values", font=("Helvetica", 8), anchor="w")
    label_list.grid(sticky="w")

    scrollbar = tk.Scrollbar(scrollbar_frame)
    scrollbar.grid(row=0, column=1, sticky='ns')
    value_list = tk.Listbox(scrollbar_frame, yscrollcommand=scrollbar.set, width=40, height=15)
    
    # Define mandatory and desirable fields
    mandatory_fields_collection = [
        'has_title',
        'was_issued',
        'was_modified',
        'has_publisher',
        'has_contributor',
        'has_creator',
        'has_owner',
        'has_responsible',
        'has_original_id',
        'has_ariadne_subject',
        'has_native_subject',
        'has_derived_subject_uri',
        'has_derived_subject_term',
        'has_language',
        'was_created_on',
        'has_access_rights',
        'has_native_period'
    ]

    mandatory_fields_individual_data_resource = [
        'has_title',
        'was_issued',
        'was_modified',
        'has_publisher',
        'has_contributor',
        'has_creator',
        'has_owner',
        'has_responsible',
        'has_original_id',
        'has_ariadne_subject',
        'has_native_subject',
        'has_derived_subject_uri',
        'has_derived_subject_term',
        'has_language',
        'was_created_on',
        'has_access_rights',
        'has_native_period',
        'has_data_type'
    ]

    desirable_fields = [
        'has_description',
        'has_access_policy',
        'has_period',
        'has_chronontology_uri',
        'from',
        'until',
        'is_part_of'
    ]
    
    def check_missing_values():
        """Check which mandatory and desirable values are missing"""
        missing_items = []
        
        # Get current values from the DataFrame
        if df_current_data is not None:
            current_properties = df_current_data['Metadata Property'].tolist()
            current_values = dict(zip(df_current_data['Metadata Property'], df_current_data['Metadata Value']))
            
            #Check if Resource_type is collection or individual data resource and choose mandatory fields accordingly
            if current_values.get('Resource_Type') == 'collection':
                mandatory_fields = mandatory_fields_collection
            else:
                mandatory_fields = mandatory_fields_individual_data_resource
            # Check mandatory fields    
            for field in mandatory_fields:
                if field not in current_properties or pd.isna(current_values.get(field)) or str(current_values.get(field)).strip() == 'not_defined':
                    missing_items.append(f"🔴 MANDATORY: {field}")
            
            # Check desirable fields
            for field in desirable_fields:
                if field not in current_properties or pd.isna(current_values.get(field)) or str(current_values.get(field)).strip() == 'not_defined':
                    missing_items.append(f"🟡 DESIRABLE: {field}")
        
        return missing_items
    
    def update_missing_values_list():
        """Update the listbox with missing values"""
        value_list.delete(0, tk.END)
        missing_items = check_missing_values()
        
        if missing_items:
            for item in missing_items:
                value_list.insert(tk.END, item)
        else:
            value_list.insert(tk.END, "✅ All mandatory fields completed!")
            value_list.insert(tk.END, "✅ All desirable fields completed!")
        
        # Update the count in the label
        mandatory_missing = len([item for item in missing_items if "MANDATORY" in item])
        desirable_missing = len([item for item in missing_items if "DESIRABLE" in item])
        label_list.config(text=f"Missing Values (🔴{mandatory_missing} mandatory, 🟡{desirable_missing} desirable)")
    
    def on_missing_value_select(event):
        """Handle selection of missing value to jump to it in the tree"""
        selection = value_list.curselection()
        if selection:
            selected_text = value_list.get(selection[0])
            # Extract field name from the text (remove emoji and prefix)
            if ":" in selected_text:
                field_name = selected_text.split(": ")[1]
                
                # Find and select the corresponding item in the tree
                for item in tree.get_children():
                    values = tree.item(item, 'values')
                    if values[0] == field_name:
                        tree.selection_set(item)
                        tree.focus(item)
                        tree.see(item)
                        break

    value_list.grid(row=0, column=0, sticky='nsew')
    scrollbar.config(command=value_list.yview)
    value_list.bind('<<ListboxSelect>>', on_missing_value_select)

    #Data Viewer in Spalte 2

    viewer_frame = tk.Frame(newdatamenu)
    viewer_frame.grid(column=1, row=2, rowspan=6, pady=10, padx=15, sticky="nsew")
    # Allow the viewer_frame to expand
    newdatamenu.grid_columnconfigure(1, weight=3)
    newdatamenu.grid_rowconfigure(2, weight=3)

    tree = ttk.Treeview(viewer_frame, columns=('Metadata Property', 'Metadata Value'), show='headings')
    tree.heading('Metadata Property', text='Metadata Property')
    tree.heading('Metadata Value', text='Metadata Value')

    # Make columns wider for better editing
    tree.column('Metadata Property', width=200)
    tree.column('Metadata Value', width=300)

    # Populate the tree if we have data
    if df_current_data is not None:
        for index, row in df_current_data.iterrows():
            tree.insert('', tk.END, values=(row['Metadata Property'], row['Metadata Value']))
    else:
        tree.insert('', tk.END, values=("No data", "Please create a dataset first"))

    # Variables for editing
    edit_item = None
    edit_entry = None

    def on_double_click(event):
        """Handle double-click to edit cell"""
        nonlocal edit_item, edit_entry
        
        # Get the selected item and column
        item = tree.selection()[0] if tree.selection() else None
        if not item:
            return
            
        # Get column clicked
        column = tree.identify_column(event.x)
        if column == '#2':  # Only allow editing the 'Metadata Value' column
            # Get the bounding box of the cell
            bbox = tree.bbox(item, column)
            if not bbox:
                return
                
            # Store current editing info
            edit_item = item
            
            # Get current value and property name
            current_value = tree.item(item, 'values')[1]
            property_name = tree.item(item, 'values')[0]
            
            # Check if this is the "is_part_of" field
            if property_name == 'is_part_of':
                # Get list of available datasets from metadata_mirror folder
                metadata_mirror_path = os.path.join(os.path.dirname(__file__), "../../metadata_mirror")
                available_datasets = []
                if os.path.exists(metadata_mirror_path):
                    xml_files = [f for f in os.listdir(metadata_mirror_path) if f.endswith('.xml')]
                    available_datasets = [os.path.splitext(f)[0] for f in xml_files]
                
                # Create combobox widget for editing
                edit_entry = ttk.Combobox(tree, values=available_datasets)
                edit_entry.place(x=bbox[0], y=bbox[1], width=bbox[2], height=bbox[3])
                
                # Set current value
                edit_entry.set(current_value)
                edit_entry.focus()
                
                # Bind events
                edit_entry.bind('<Escape>', cancel_edit)
                edit_entry.bind('<Return>', save_edit)
                edit_entry.bind('<FocusOut>', save_edit)
            else:
                # Create normal entry widget for editing
                edit_entry = tk.Entry(tree)
                edit_entry.place(x=bbox[0], y=bbox[1], width=bbox[2], height=bbox[3])
                
                # Set current value
                edit_entry.insert(0, current_value)
                edit_entry.select_range(0, tk.END)
                edit_entry.focus()
                
                # Bind events
                edit_entry.bind('<Escape>', cancel_edit)
                edit_entry.bind('<Return>', save_edit)
                #End also when clicking outside
                edit_entry.bind('<FocusOut>', save_edit)

    def cancel_edit(event=None):
        """Cancel editing without saving"""
        cleanup_edit()

    def cleanup_edit():
        """Clean up editing widgets"""
        nonlocal edit_item, edit_entry
        
        if edit_entry:
            edit_entry.destroy()
            edit_entry = None
        edit_item = None

    # Bind double-click event
    tree.bind('<Double-1>', on_double_click)

    tree.pack(pady=10, padx=15, fill=tk.BOTH, expand=True)

    # Add instructions
    instructions_label = tk.Label(viewer_frame, 
                                 text="Double-click on 'Metadata Value' cells to edit. Press Escape to cancel and button to save.",
                                 font=("Helvetica", 8), fg="gray")
    instructions_label.pack(pady=(0, 5))
    
    # Initialize the missing values list
    update_missing_values_list()

    def open_documentation():
        """Open the documentation in a web browser"""
        #The file is /documentation/MTT_Readme.html
        doc_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../documentation/MTT_Readme.html"))
        webbrowser.open_new(f"file:///{doc_path.replace(os.sep, '/')}")



   # Back button
    back_button = tk.Button(newdatamenu, text="Back", command=go_back, width=25, height=1, bg="orange", fg="white", font=("Helvetica", 10, "bold"))
    back_button.grid(row=9, column=0, columnspan=2, padx=10, pady=5, sticky="w")

    # Open documentation button
    open_doc_button = tk.Button(newdatamenu, text="Open Documentation", command=open_documentation, width=25, height=1, bg="light blue", fg="white", font=("Helvetica", 10, "bold"))
    open_doc_button.grid(row=9, column=0, columnspan=2, padx=10, pady=5)

    # Cancel button
    cancel_button = tk.Button(newdatamenu, text="Cancel without Saving", command=go_back, width=25, height=1, bg="red", fg="black", font=("Helvetica", 10, "bold"))
    cancel_button.grid(row=9, column=0, columnspan=2, padx=10, pady=5, sticky="e")

    

if __name__ == "__main__":
    import tkinter as tk
    from tkinter import messagebox
    from tkinter import ttk
    from datetime import datetime
    import pandas as pd
    import os
    import sys
    import importlib
    import webbrowser
    import xml.dom.minidom
    sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
    from recognize_repository import rec_repo
    import xml_conversion
    new_dataset()
    tk.mainloop()
else:
    import tkinter as tk
    from tkinter import messagebox
    from tkinter import ttk
    from datetime import datetime
    import pandas as pd
    import os
    import sys
    import importlib
    import webbrowser
    import xml.dom.minidom
    sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
    from recognize_repository import rec_repo
    import xml_conversion
    from modules.menu import main_menu
