import os

def set_window_icon(window):
    """Helper function to set the custom icon for any tkinter window"""
    try:
        icon_path = os.path.join(os.path.dirname(__file__), "MTT_favicon.png")
        if os.path.exists(icon_path):
            from PIL import Image, ImageTk
            icon_image = Image.open(icon_path)
            icon_photo = ImageTk.PhotoImage(icon_image)
            window.iconphoto(True, icon_photo)
    except Exception as e:
        print(f"Could not load icon: {e}")

def main_menu_design(root=None):
    # Use provided root or create new one
    if root is None:
        mainmenu = tk.Tk()
    else:
        mainmenu = root
        
    mainmenu.title("FAIR.rdm MTT")
    mainmenu.geometry("800x750")
    mainmenu.configure(bg="#f0f0f0")
    
    # Set custom icon
    set_window_icon(mainmenu)

    def button1_action():
        print("Button 1 clicked")
        #Execute module named person-institution.py in modules/menu
        new_person_and_institution.person_institution()
        #Close the main window
        mainmenu.destroy()

    def button2_action():
        print("Button 2 clicked")
        open_dialog(mainmenu)

    def open_dialog(parent):
        """
        Öffnet ein neues Top-Level-Fenster (Dialogfenster), um den Datensatz-Typ abzufragen.
        """
        dialog = tk.Toplevel(parent)
        dialog.title("Initial Input")
        dialog.geometry("300x250")
        
        # Zentriert das Dialogfenster über dem Hauptfenster
        parent.update_idletasks()
        main_x = parent.winfo_x()
        main_y = parent.winfo_y()
        main_width = parent.winfo_width()
        main_height = parent.winfo_height()
        dialog_width = dialog.winfo_width()
        dialog_height = dialog.winfo_height()
        
        dialog.geometry(
            f"+{main_x + (main_width - dialog_width) // 2}"
            f"+{main_y + (main_height - dialog_height) // 2}"
        )

        # Stellt sicher, dass das Hauptfenster nicht bedienbar ist,
        # solange der Dialog offen ist
        dialog.transient(parent)
        dialog.grab_set()

        # UI-Elemente im Dialogfenster
        label = ttk.Label(dialog, text="Please choose the suitable type:")
        label.pack(pady=10)

        radio_frame = ttk.Frame(dialog)
        radio_frame.pack(pady=5)

        # Variable für die Radiobuttons (specify dialog as master)
        choice_var = tk.StringVar(master=dialog, value="nd")

        # Radiobuttons für die Auswahl des Datensatz-Typs
        radio_collection = ttk.Radiobutton(
            radio_frame,
            text="Collection",
            value="collection",
            variable=choice_var,
        )
        radio_collection.pack(side="left", padx=10)

        radio_individual = ttk.Radiobutton(
            radio_frame,
            text="Individual Data Resource",
            value="individual",
            variable=choice_var,
        )
        radio_individual.pack(side="left", padx=10)

        # Print the selected dataset type whenever a radio button is selected
        def update_dataset_type(*args):
            print(f"Selected dataset type: {choice_var.get()}")
            dataset_type = choice_var.get()
            return dataset_type
        choice_var.trace_add("write", update_dataset_type)
        print(f"Initial dataset type: {choice_var.get()}")
        
        # Load suggestions from metadata_tables directory
        metadata_tables_path = os.path.join(os.path.dirname(__file__), "../../metadata_tables")
        name_suggestions = []
        
        try:
            if os.path.exists(metadata_tables_path):
                metadata_files = [f for f in os.listdir(metadata_tables_path) if os.path.isfile(os.path.join(metadata_tables_path, f))]
                for file in metadata_files:
                    #if file.startswith('P') or file.startswith('S'):
                        # Remove file extension and add to suggestions
                        name_without_extension = os.path.splitext(file)[0]
                        name_suggestions.append(name_without_extension)
        except Exception as e:
            print(f"Could not load metadata files: {e}")
            # Fallback to default suggestions if directory can't be read
            name_suggestions = [
                "Something's wrong",
                "This is probably an error", 
                "Folder can't be read or is missing"
            ]
    
        #Enter the datasets name with autocomplete
        label_name = ttk.Label(dialog, text="Dataset Name:")
        label_name.pack(pady=5)
        
        # Create frame for entry and suggestions
        name_frame = ttk.Frame(dialog)
        name_frame.pack(pady=5, padx=10, fill=tk.X)
        
        dialog.entry_name = ttk.Entry(name_frame)
        dialog.entry_name.pack(fill=tk.X)
        
        # Create listbox for suggestions (initially hidden)
        suggestions_listbox = tk.Listbox(name_frame, height=5)
        suggestions_listbox.pack(fill=tk.X)
        suggestions_listbox.pack_forget()  # Hide initially
        
        def on_name_keyup(event):
            """Filter suggestions based on user input"""
            current_text = dialog.entry_name.get().lower()
            
            if current_text:
                # Filter suggestions that contain the current text
                filtered_suggestions = [s for s in name_suggestions if current_text in s.lower()]
                
                if filtered_suggestions:
                    suggestions_listbox.delete(0, tk.END)
                    for suggestion in filtered_suggestions[:5]:  # Show max 5 suggestions
                        suggestions_listbox.insert(tk.END, suggestion)
                    suggestions_listbox.pack(fill=tk.X)
                else:
                    suggestions_listbox.pack_forget()
            else:
                suggestions_listbox.pack_forget()
        
        def on_suggestion_select(event):
            """Handle suggestion selection"""
            selection = suggestions_listbox.curselection()
            if selection:
                selected_text = suggestions_listbox.get(selection[0])
                dialog.entry_name.delete(0, tk.END)
                dialog.entry_name.insert(0, selected_text)
                suggestions_listbox.pack_forget()
        
        def hide_suggestions(event):
            """Hide suggestions when clicking outside"""
            suggestions_listbox.pack_forget()
        
        # Bind events
        dialog.entry_name.bind('<KeyRelease>', on_name_keyup)
        suggestions_listbox.bind('<Double-Button-1>', on_suggestion_select)
        dialog.bind('<Button-1>', hide_suggestions)
        
        #Enter the datasets uri
        label_uri = ttk.Label(dialog, text="Dataset URI:")
        label_uri.pack(pady=5)
        dialog.entry_uri = ttk.Entry(dialog)
        dialog.entry_uri.pack(pady=5, padx=10, fill=tk.X)
        # Default value for the entry fields
        dialog.entry_uri.insert(0, "blank")

        ok_button = ttk.Button(dialog, text="OK", command=lambda: on_dialog_close(choice_var.get(), dialog.entry_name.get(), dialog.entry_uri.get(), name_suggestions))
        ok_button.pack(pady=10)

        # Warte, bis das Dialogfenster geschlossen wird
        #parent.wait_window(dialog)


    def on_dialog_close(dataset_type, dataset_name, dataset_uri, name_suggestions): 
        """
        Wird aufgerufen, wenn der OK-Button im Dialogfenster gedrückt wird.
        Holt die Auswahl und schließt das Dialogfenster.
        """
        print(f"Type: {dataset_type}")
        print(f"Dataset Name: {dataset_name}")
        print(f"Dataset URI: {dataset_uri}")

        #dataset_name has to be new. It cannot be an existing name like in suggestions
        if dataset_name in name_suggestions:
            messagebox.showerror("Error", "Dataset name already exists. Please choose a different name.")
            return

        if dataset_type == "collection":
            print("Collection selected")
            #Create a copy of SPP2143_ARIADNE_Collection_Import_Template.xlsx from templates in the metadata_tables directory
            #name it with the dataset_name and save it in the metadata_tables directory
            metadata_tables_path = os.path.join(os.path.dirname(__file__), "../../metadata_tables")
            template_path = os.path.join(os.path.dirname(__file__), "../../templates", "SPP2143_ARIADNE_Collection_Import_Template.xlsx")
            shutil.copy(template_path, os.path.join(metadata_tables_path, f"{dataset_name}.xlsx"))
        else:
            print("Individual Data Resource selected")
            #Copy of SPP2143_ARIADNE_IDR_Import_Template.xlsx
            metadata_tables_path = os.path.join(os.path.dirname(__file__), "../../metadata_tables")
            template_path = os.path.join(os.path.dirname(__file__), "../../templates", "SPP2143_ARIADNE_IDR_Import_Template.xlsx")
            shutil.copy(template_path, os.path.join(metadata_tables_path, f"{dataset_name}.xlsx"))
        
        if dataset_uri == "blank":
            print("loading empty template")
            #Read in the template and fill in the dataset_name
            wb = load_workbook(os.path.join(metadata_tables_path, f"{dataset_name}.xlsx"))
            ws = wb['metadata']
            # Find the row with 'has_identifier' and update column B (2nd column)
            for row in ws.iter_rows():
                if row[0].value == 'has_identifier':
                    row[1].value = dataset_name
                    # Optionally add formatting to the new value
                    row[1].font = Font(bold=True)
                    break
            #Insert current date in "was_issued"
            for row in ws.iter_rows():
                if row[0].value == 'was_issued':
                    current_date = datetime.now().strftime("%Y-%m-%d")
                    row[1].value = current_date
                    break
            #And current date in "was_modified"
            for row in ws.iter_rows():
                if row[0].value == 'was_modified':
                    current_date = datetime.now().strftime("%Y-%m-%d")
                    row[1].value = current_date
                    break
            # Save the file with preserved formatting
            wb.save(os.path.join(metadata_tables_path, f"{dataset_name}.xlsx"))
            print("File saved with preserved formatting!")

        else:
            print(f"Preparing Template for: {dataset_uri}")
            #Read in the template and fill in the dataset_name and dataset_uri
            wb = load_workbook(os.path.join(metadata_tables_path, f"{dataset_name}.xlsx"))
            ws = wb['metadata']

            # Find the row with 'has_identifier' and update column B (2nd column)
            for row in ws.iter_rows():
                if row[0].value == 'has_identifier':
                    row[1].value = dataset_name
                    # Optionally add formatting to the new value
                    row[1].font = Font(bold=True)
                    break
            for row in ws.iter_rows():
                if row[0].value == 'has_landing_page':
                    row[1].value = dataset_uri
                    break
            #Insert current date in "was_issued"
            for row in ws.iter_rows():
                if row[0].value == 'was_issued':
                    current_date = datetime.now().strftime("%Y-%m-%d")
                    row[1].value = current_date
                    break
            #And current date in "was_modified"
            for row in ws.iter_rows():
                if row[0].value == 'was_modified':
                    current_date = datetime.now().strftime("%Y-%m-%d")
                    row[1].value = current_date
                    break
            
            # Save the file with preserved formatting
            wb.save(os.path.join(metadata_tables_path, f"{dataset_name}.xlsx"))
            print("File saved with preserved formatting!")

        #Insert the identifier into the log.xlsx with status "created"
        log_xlsx = os.path.join(metadata_tables_path, "log.xlsx")
        try:
            if os.path.exists(log_xlsx):
                wb_log = load_workbook(log_xlsx)
                ws_log = wb_log.active

                if ws_log.max_row == 1 and (ws_log['A1'].value is None or str(ws_log['A1'].value).strip() != 'identifier'):
                    ws_log.cell(row=1, column=1, value='identifier')
                    ws_log.cell(row=1, column=2, value='metadata_status')

                next_row = ws_log.max_row + 1
                ws_log.cell(row=next_row, column=1, value=dataset_name)
                ws_log.cell(row=next_row, column=2, value='created')
                wb_log.save(log_xlsx)
                print("Log updated with new dataset entry.")
        except Exception as e:
            print(f"Could not update log.xlsx: {e}")

        mainmenu.destroy()  # Schließt das Hauptfenster
        new_dataset.new_dataset()
    
    ###

    def button3_action():
        '''
        Open the menu to update or delete an existing dataset
        '''
        print("Button 3 clicked")
        #Get the file currently selected in the value_list Listbox
        selected_file = None
        try:
            selection = value_list.curselection()
            if not selection:
                messagebox.showerror("Error", "Please select a dataset from the Data Selector list.")
                return
            selected_display = value_list.get(selection[0])
            selected_file = display_to_file.get(selected_display, selected_display)
            print(f"Selected file: {selected_file}")
        except Exception as e:
            messagebox.showerror("Error", "Please select a dataset from the Data Selector list.")
            return
        mainmenu.destroy()
        change_dataset.change_dataset(selected_file=selected_file)

    def button4_action():
        '''
        Sync the content of metadata_mirror to GitLab
        '''
        print("Button 4 clicked")
        #Execute the sync function from the sync module
        metadata_mirror_path = os.path.join(os.path.dirname(__file__), f"../../{mtt_config.local_folder}")
        sync.mirror_to_gitlab(metadata_mirror_path, mtt_config.repo_url, mtt_config.target_subdir, mtt_config.token, mtt_config.branch)
        #Show a messagebox while syncing is in progress
        #get a list of all files in the metadata_mirror
        mirror_files = []
        try:
            mirror_files = [f for f in os.listdir(metadata_mirror_path) if os.path.isfile(os.path.join(metadata_mirror_path, f))]
        except Exception as e:
            print(f"Could not load metadata mirror files: {e}")
        #remove the file extension from the filenames in mirror_files
        mirror_files = [os.path.splitext(f)[0] for f in mirror_files]
        #set status in log.xlsx to "mirrored" for all files in the metadata_mirror
        log_xlsx = os.path.join(os.path.dirname(__file__), "../../metadata_tables/log.xlsx")
        try:
            if os.path.exists(log_xlsx):
                wb_log = load_workbook(log_xlsx)
                ws_log = wb_log['Sheet1']
                for row in ws_log.iter_rows(min_row=2):  # Assuming the first row is a header
                    filename_cell = row[0]  # Assuming the filename is in the first column
                    status_cell = row[1]    # Assuming the status is in the second column
                    if filename_cell.value in mirror_files:
                        status_cell.value = "mirrored"
                wb_log.save(log_xlsx)
                print("Log updated with mirrored status.")
        except Exception as e:
            print(f"Could not update log.xlsx: {e}")
        #refresh the value_list to show the updated status icons
        value_list.delete(0, tk.END)
        #List all files in metadata_tables with status icon
        metadata_tables_path = os.path.join(os.path.dirname(__file__), "../../metadata_tables")
        status_by_identifier = {}
        try:
            log_xlsx = os.path.join(metadata_tables_path, "log.xlsx")
            log_csv = os.path.join(metadata_tables_path, "log.csv")
            df_log = None
            if os.path.exists(log_xlsx):
                try:
                    df_log = pd.read_excel(log_xlsx, sheet_name="Sheet1")
                except Exception:
                    df_log = pd.read_excel(log_xlsx)
            elif os.path.exists(log_csv):
                df_log = pd.read_csv(log_csv)
            if df_log is not None and 'identifier' in df_log.columns and 'metadata_status' in df_log.columns:
                status_by_identifier = {str(row['identifier']).strip(): str(row['metadata_status']).strip() for _, row in df_log.iterrows()}
        except Exception as e:
            print(f"Could not load df_log: {e}")
        metadata_files = [f for f in os.listdir(metadata_tables_path) if os.path.isfile(os.path.join(metadata_tables_path, f))]
        status_icon = {
            'created': '❌',
            'completed': '⭕',
            'converted': '🟣',
            'mirrored': '✔️'
        }
        for file in metadata_files:
            if "log" in file or "registered_persons" in file or not file.endswith('.xlsx'):
                continue
            identifier_no_ext = os.path.splitext(file)[0]
            status = status_by_identifier.get(file) or status_by_identifier.get(identifier_no_ext)
            icon = status_icon.get(str(status).lower(), '') if status else ''
            display_text = f"{icon} {file}" if icon else file
            value_list.insert(tk.END, display_text)
            display_to_file[display_text] = file

    def open_metadata_mirror():
        '''
        Open the metadata_mirror folder in the default file explorer
        '''
        print("Open Metadata Mirror clicked")
        #Open the metadata_mirror folder in the default file explorer
        metadata_mirror_path = os.path.join(os.path.dirname(__file__), f"../../{mtt_config.local_folder}")
        if os.path.exists(metadata_mirror_path):
            webbrowser.open(metadata_mirror_path)
        else:
            messagebox.showerror("Error", f"Metadata mirror folder not found: {metadata_mirror_path}")

    def open_metadata_tables():
        '''
        Open the metadata_tables folder in the default file explorer
        '''
        print("Open Metadata Tables clicked")
        #Open the metadata_tables folder in the default file explorer
        metadata_tables_path = os.path.join(os.path.dirname(__file__), "../../metadata_tables")
        if os.path.exists(metadata_tables_path):
            webbrowser.open(metadata_tables_path)
        else:
            messagebox.showerror("Error", f"Metadata tables folder not found: {metadata_tables_path}")

    def open_config():
        """Open the configuration file in the default text editor"""
        config_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "../mtt_config.py"))
        try:
            if os.name == 'nt':  # For Windows
                os.startfile(config_path)
        except AttributeError:
            messagebox.showerror("Error", f"Could not open configuration file.")

    def open_registered_persons():
        '''
        Open the registered_persons.xlsx file in the default application
        '''
        print("Open Registered Persons clicked")
        #Open the registered_persons.xlsx file in the default application
        registered_persons_path = os.path.join(os.path.dirname(__file__), "../../metadata_tables/registered_persons.xlsx")
        if os.path.exists(registered_persons_path):
            webbrowser.open(registered_persons_path)
        else:
            messagebox.showerror("Error", f"Registered persons file not found: {registered_persons_path}")

    def open_getty_aat():
        '''
        Open the Getty AAT website in the default web browser
        '''
        print("Open Getty AAT clicked")
        webbrowser.open_new_tab(mtt_config.getty_aat)

    def open_ariadne_portal():
        '''
        Open the ARIADNE Portal website in the default web browser
        '''
        print("Open ARIADNE Portal clicked")
        webbrowser.open_new_tab(mtt_config.ariadne_portal)

    def open_staging_portal():
        '''
        Open the Staging Portal website in the default web browser
        '''
        print("Open Staging Portal clicked")
        webbrowser.open_new_tab(mtt_config.staging_portal)

    def open_staging_graph_db():
        '''
        Open the Staging Graph DB website in the default web browser
        '''
        print("Open Staging Graph DB clicked")
        webbrowser.open_new_tab(mtt_config.staging_graph_db)

    def open_chronontology():
        '''
        Open the iDAI.chronontology website in the default web browser
        '''
        print("Open iDAI.chronontology clicked")
        webbrowser.open_new_tab(mtt_config.chronontology)

    def open_periodo():
        '''
        Open the PeriodO website in the default web browser
        '''
        print("Open PeriodO clicked")
        webbrowser.open_new_tab(mtt_config.periodo)

    def open_ao_cat():
        '''
        Open the AO Cat website in the default web browser
        '''
        print("Open AO Cat clicked")
        webbrowser.open_new_tab(mtt_config.ao_cat)

    def open_lexvo():
        '''
        Open the Lexvo website in the default web browser
        '''
        print("Open Lexvo clicked")
        webbrowser.open_new_tab(mtt_config.lexvo)

    def open_gitlab_repo():
        '''
        Open the GitLab repository in the default web browser
        '''
        print("Open GitLab Repo clicked")
        webbrowser.open_new_tab(mtt_config.git_repo)

    def open_joai():
        '''
        Open the jOAI website in the default web browser
        '''
        print("Open jOAI clicked")
        webbrowser.open_new_tab(mtt_config.oai_pmh_status)
    
    def open_oai_pmh():
        '''
        Open the OAI-PMH website in the default web browser
        '''
        print("Open OAI-PMH clicked")
        webbrowser.open_new_tab(mtt_config.oai_pmh_list)

    def open_3m():
        '''
        Open the 3M website in the default web browser
        '''
        print("Open 3M clicked")
        webbrowser.open_new_tab(mtt_config.three_m)

    def open_github():
        '''
        Open the GitHub website in the default web browser
        '''
        print("Open GitHub clicked")
        webbrowser.open_new_tab(mtt_config.git_hub)

    def open_documentation():
        """Open the documentation in a web browser"""
        #The file is /documentation/MTT_Readme.html
        doc_path2 = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../documentation/MTT_Readme.html"))
        webbrowser.open(f"file:///{doc_path2.replace(os.sep, '/')}")

    def open_metadata_standard():
        """Open the SPP-2143 Metadata Standard in a web browser"""
        doc_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../documentation/SPP_2143_Metadata_Standard_Documentation.html"))
        webbrowser.open(f"file:///{doc_path.replace(os.sep, '/')}")

    mainmenu.grid_columnconfigure(0, weight=1)
    mainmenu.grid_columnconfigure(1, weight=1)
    mainmenu.grid_rowconfigure(0, weight=1)
    mainmenu.grid_rowconfigure(1, weight=1)
    mainmenu.grid_rowconfigure(2, weight=1)
    mainmenu.grid_rowconfigure(3, weight=1)
    mainmenu.grid_rowconfigure(4, weight=1)
    mainmenu.grid_rowconfigure(5, weight=1)
    mainmenu.grid_rowconfigure(6, weight=1)
    mainmenu.grid_rowconfigure(7, weight=1)
    mainmenu.grid_rowconfigure(8, weight=1)

    #Titel 1. Zeile
    label_title = tk.Label(mainmenu, text="FAIR.rdm Metadata Tables Tool", font=("Helvetica", 16, "bold"), anchor="w")
    label_title.grid(row=0, column=0, columnspan=2, pady=(5, 0), padx=10, sticky="nw")

    #Untertitel 2. Zeile
    label_subtitle = tk.Label(mainmenu, text="Metadata Converter from multiple repositories to ARIADNE Portal", font=("Helvetica", 12, "bold"), anchor="w")
    label_subtitle.grid(row=1, column=0, columnspan=2, pady=(3, 0), padx=10, sticky="nw")

    menu_bar = tk.Menu(mainmenu)

    data_menu = tk.Menu(menu_bar, tearoff=0)
    data_menu.add_command(label="Change Dataset", command=button3_action, accelerator="Ctrl+C")
    data_menu.add_command(label="New Data Record", command=button2_action, accelerator="Ctrl+N")
    data_menu.add_command(label="New Person / Institution", command=button1_action, accelerator="Ctrl+P")
    data_menu.add_command(label="Push to GitLab", command=button4_action, accelerator="Ctrl+G")
    data_menu.add_separator()
    data_menu.add_command(label="Retrieve Data", accelerator="Ctrl+R", state="disabled")
    data_menu.add_command(label="Save Progress", accelerator="Ctrl+S", state="disabled")
    data_menu.add_command(label="Convert to XML", accelerator="Ctrl+X", state="disabled")
    data_menu.add_separator()
    data_menu.add_command(label="Delete Dataset", state="disabled")
    menu_bar.add_cascade(label="Data", menu=data_menu)

    open_menu = tk.Menu(menu_bar, tearoff=0)
    open_menu.add_command(label="Metadata Mirror", command=open_metadata_mirror)
    open_menu.add_command(label="Metadata Tables", command=open_metadata_tables)
    open_menu.add_command(label="Config File", command=open_config)
    open_menu.add_command(label="Registered Persons", command=open_registered_persons)
    open_menu.add_separator()
    open_menu.add_command(label="Open with Excel", state="disabled")
    open_menu.add_command(label="Show XML", state="disabled")
    menu_bar.add_cascade(label="Open", menu=open_menu)

    links_menu = tk.Menu(menu_bar, tearoff=0)
    links_menu.add_command(label="Getty AAT", command=open_getty_aat)
    links_menu.add_command(label="ARIADNE Portal", command=open_ariadne_portal)
    links_menu.add_command(label="Staging Portal", command=open_staging_portal)
    links_menu.add_command(label="Staging Graph DB", command=open_staging_graph_db)
    links_menu.add_command(label="iDAI.chronontology", command=open_chronontology)
    links_menu.add_command(label="PeriodO", command=open_periodo)
    links_menu.add_command(label="AO Cat", command=open_ao_cat)
    links_menu.add_command(label="Lexvo", command=open_lexvo)
    links_menu.add_command(label="GitLab Repo", command=open_gitlab_repo)
    links_menu.add_command(label="jOAI", command=open_joai)
    links_menu.add_command(label="OAI-PMH", command=open_oai_pmh)
    links_menu.add_command(label="3M", command=open_3m)
    links_menu.add_command(label="GitHub", command=open_github)
    menu_bar.add_cascade(label="Links", menu=links_menu)

    help_menu = tk.Menu(menu_bar, tearoff=0)
    help_menu.add_command(label="MTT Documentation", command=open_documentation)
    help_menu.add_command(label="SPP-2143-Metadata-Standard", command=open_metadata_standard)
    help_menu.add_separator()
    help_menu.add_command(label="About", state="disabled")
    menu_bar.add_cascade(label="Help", menu=help_menu)

    mainmenu.config(menu=menu_bar)

    # Keyboard shortcuts
    mainmenu.bind("<Control-c>", lambda e: button3_action())
    mainmenu.bind("<Control-n>", lambda e: button2_action())
    mainmenu.bind("<Control-p>", lambda e: button1_action())
    mainmenu.bind("<Control-g>", lambda e: button4_action())
    #mainmenu.bind("<Control-r>", lambda e: select_all())
    #mainmenu.bind("<Control-s>", lambda e: select_all())
    #mainmenu.bind("<Control-x>", lambda e: select_all())
    
    #Data Selector in Spalte 0 (in der Beta in Zeile 8)

    # Prepare status icons from df_log (identifier, status)
    metadata_tables_path = os.path.join(os.path.dirname(__file__), "../../metadata_tables")
    status_by_identifier = {}
    try:
        log_xlsx = os.path.join(metadata_tables_path, "log.xlsx")
        log_csv = os.path.join(metadata_tables_path, "log.csv")
        df_log = None
        if os.path.exists(log_xlsx):
            try:
                df_log = pd.read_excel(log_xlsx, sheet_name="Sheet1")
            except Exception:
                df_log = pd.read_excel(log_xlsx)
        elif os.path.exists(log_csv):
            df_log = pd.read_csv(log_csv)
        if df_log is not None and 'identifier' in df_log.columns and 'metadata_status' in df_log.columns:
            status_by_identifier = {str(row['identifier']).strip(): str(row['metadata_status']).strip() for _, row in df_log.iterrows()}
    except Exception as e:
        print(f"Could not load df_log: {e}")

    scrollbar_frame = tk.Frame(mainmenu)
    scrollbar_frame.grid(row=2, column=0, rowspan=6, pady=0, padx=15, sticky="nsew")
    #allow the scrollbar_frame to expand to match the height of viewer_frame
    mainmenu.grid_columnconfigure(0, weight=4)
    mainmenu.grid_rowconfigure(2, weight=4)
    scrollbar_frame.grid_rowconfigure(1, weight=1)
    scrollbar_frame.grid_columnconfigure(0, weight=1)

    label_list = tk.Label(scrollbar_frame, text="Data Selector", font=("Helvetica", 8), anchor="w")
    label_list.grid(row=0, column=0, columnspan=1, sticky="w")

    scrollbar = tk.Scrollbar(scrollbar_frame)
    scrollbar.grid(row=1, column=1, sticky='ns')
    value_list = tk.Listbox(scrollbar_frame, yscrollcommand=scrollbar.set, width=40)
    display_to_file = {}
    #for i in range(100):
        #value_list.insert(tk.END, str(i))
    # List all files in metadata_tables with status icon
    metadata_files = [f for f in os.listdir(metadata_tables_path) if os.path.isfile(os.path.join(metadata_tables_path, f))]
    status_icon = {
        'created': '❌',
        'completed': '⭕',
        'converted': '🟣',
        'mirrored': '✔️'
    }
    for file in metadata_files:
        if "log" in file or "registered_persons" in file or not file.endswith('.xlsx'):
            continue
        identifier_no_ext = os.path.splitext(file)[0]
        status = status_by_identifier.get(file) or status_by_identifier.get(identifier_no_ext)
        icon = status_icon.get(str(status).lower(), '') if status else ''
        display_text = f"{icon} {file}" if icon else file
        value_list.insert(tk.END, display_text)
        display_to_file[display_text] = file

    value_list.grid(row=1, column=0, sticky='nsew')
    scrollbar.config(command=value_list.yview)
    value_list.bind('<<ListboxSelect>>', lambda event: on_select(event))

    #Selected file in value_list to dataframe and display in the data viewer
    # This function will be called when an item in the listbox is selected
    def on_select(event):
        selection = value_list.curselection()
        if not selection:
            return
        selected_display = value_list.get(selection[0])
        selected_file = display_to_file.get(selected_display, selected_display)
        file_path = os.path.join(metadata_tables_path, selected_file)
        df = pd.read_excel(file_path, sheet_name="metadata")  # Assuming the files are Excel files
        df = df.fillna('not_defined')  # Fill NaN values with empty strings
        # Clear the treeview
        for item in tree.get_children():
            tree.delete(item)
        # Insert new data into the treeview
        for index, row in df.iterrows():
            tree.insert('', tk.END, values=(row['Metadata Property'], row['Metadata Value']))  # Adjust column names as needed
            
    #Data Viewer in Spalte 2

    viewer_frame = tk.Frame(mainmenu)
    viewer_frame.grid(column=1, row=2, rowspan=6, pady=0, padx=15, sticky="nsew")
    # Allow the viewer_frame to expand
    mainmenu.grid_columnconfigure(1, weight=3)
    mainmenu.grid_rowconfigure(2, weight=3)

    df_test = pd.DataFrame({
        'Metadata Property': ['Property1', 'Property2', 'Property3'],
        'Metadata Value': ['Value1', 'Value2', 'Value3']
    })

    tree = ttk.Treeview(viewer_frame, columns=('Metadata Property', 'Metadata Value'), show='headings')
    tree.heading('Metadata Property', text='Metadata Property')
    tree.heading('Metadata Value', text='Metadata Value')

    for index, row in df_test.iterrows():
        tree.insert('', tk.END, values=(row['Metadata Property'], row['Metadata Value']))

    tree.pack(pady=10, padx=15, fill=tk.BOTH, expand=True)



    # Exit button
    exit_button = tk.Button(mainmenu, text="Exit", command=mainmenu.quit, width=25, height=1, bg="red", fg="white", font=("Helvetica", 10, "bold"))
    exit_button.grid(row=8, column=0, columnspan=2, padx=10, pady=5, sticky="w")

if __name__ == "__main__":
    import tkinter as tk
    from tkinter import ttk
    from tkinter import messagebox
    import pandas as pd
    import os
    import subprocess
    import tempfile
    from datetime import datetime
    import shutil
    import webbrowser
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    main_menu_design()
    # Start the main loop
    tk.mainloop()
else:
    import tkinter as tk
    from tkinter import ttk
    from tkinter import messagebox
    import pandas as pd
    from modules.menu import new_person_and_institution
    from modules.menu import new_dataset
    from modules.menu import change_dataset
    from modules import mtt_config
    from modules import sync
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    import os
    import subprocess
    import tempfile
    from datetime import datetime
    import shutil
    import webbrowser


